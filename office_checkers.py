import os
from pathlib import Path
import docx
import fitz  # PyMuPDF
import openpyxl

# Разрешенные ГОСТовские шрифты
ALLOWED_FONTS = ["Times New Roman", "Arial", "GOST type A", "GOST type B", "ISOCPEUR"]

def check_docx(file_path: str) -> dict:
    """Анализирует DOCX файл на соответствие ГОСТ 2.105-2019"""
    errors = []
    warnings = []
    try:
        doc = docx.Document(file_path)
        
        # Проверяем стили
        for style in doc.styles:
            if style.type == docx.enum.style.WD_STYLE_TYPE.PARAGRAPH:
                font = style.font
                if font.name and font.name not in ALLOWED_FONTS:
                    warnings.append(f"В стиле '{style.name}' используется не ГОСТ шрифт: {font.name}")
        
        # Проверяем параграфы
        for i, para in enumerate(doc.paragraphs):
            # Проверка шрифта в прогонах (runs)
            for run in para.runs:
                if run.font.name and run.font.name not in ALLOWED_FONTS:
                    errors.append(f"Абзац {i+1}: недопустимый шрифт '{run.font.name}'")
                    break # Одного предупреждения на абзац достаточно
                    
            # Проверка выравнивания (по ГОСТ обычно выравнивание по ширине - JUSTIFY, но оставим как предупреждение)
            # В docx выравнивание может быть None (наследуется от стиля)
            if para.alignment == docx.enum.text.WD_ALIGN_PARAGRAPH.LEFT:
                warnings.append(f"Абзац {i+1}: выравнивание по левому краю (ГОСТ требует по ширине)")
                
    except Exception as e:
        return {"error": f"Ошибка чтения DOCX: {str(e)}"}
        
    return {
        "passed": len(errors) == 0,
        "total_errors": len(errors),
        "total_warnings": len(warnings),
        "errors": errors[:50], # Ограничиваем вывод
        "warnings": warnings[:50]
    }

def check_xlsx(file_path: str) -> dict:
    """Анализирует XLSX файл"""
    errors = []
    warnings = []
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            # Проверяем первые 50 строк и 20 столбцов для скорости
            for row in sheet.iter_rows(min_row=1, max_row=50, min_col=1, max_col=20):
                for cell in row:
                    if cell.font and cell.font.name:
                        if cell.font.name not in ALLOWED_FONTS and cell.font.name != "Calibri": # В Excel часто Calibri по умолчанию
                            warnings.append(f"Лист '{sheet_name}', ячейка {cell.coordinate}: нестандартный шрифт '{cell.font.name}'")
    except Exception as e:
        return {"error": f"Ошибка чтения XLSX: {str(e)}"}
        
    # Дедубликация предупреждений (так как их может быть тысячи)
    warnings = list(set(warnings))
    
    return {
        "passed": len(errors) == 0,
        "total_errors": len(errors),
        "total_warnings": len(warnings),
        "errors": errors[:50],
        "warnings": warnings[:50]
    }

def check_pdf(file_path: str) -> dict:
    """Анализирует векторный PDF файл (шрифты)"""
    errors = []
    warnings = []
    try:
        doc = fitz.open(file_path)
        fonts_used = set()
        for page_num in range(len(doc)):
            page = doc.load_page(page_num)
            font_list = page.get_fonts()
            for f in font_list:
                # f[3] - это базовое имя шрифта (например 'Arial-BoldMT')
                font_name = f[3].split('+')[-1] # Убираем подмножества типа ABCDEF+
                fonts_used.add(font_name)
                
        # Проверяем собранные шрифты
        for font in fonts_used:
            is_allowed = False
            for allowed in ALLOWED_FONTS:
                if allowed.lower() in font.lower() or "gost" in font.lower():
                    is_allowed = True
                    break
            if not is_allowed and "symbol" not in font.lower() and "zapf" not in font.lower():
                errors.append(f"В PDF обнаружен нестандартный шрифт: {font}")
                
    except Exception as e:
        return {"error": f"Ошибка чтения PDF: {str(e)}"}
        
    return {
        "passed": len(errors) == 0,
        "total_errors": len(errors),
        "total_warnings": len(warnings),
        "errors": errors[:50],
        "warnings": warnings[:50]
    }
